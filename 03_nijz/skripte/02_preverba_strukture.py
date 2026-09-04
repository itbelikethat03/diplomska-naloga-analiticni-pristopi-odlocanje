# -*- coding: utf-8 -*-
"""
Faza 2, korak 2: preverba predpostavk o strukturi tedenskih porocil (nacrt 2.1).

Za vsako .xlsx datoteko v data/raw/ se preveri:
- imena listov,
- glava lista 'Tb 01' (nacionalni agregat po VZS) in stevilo stolpcev,
- stevilo podatkovnih vrstic (stevilo VZS),
- datum porocila iz imena datoteke.

Rezultat: data/struktura_pregled.csv + povzetek razlicic glave na stdout.
"""

import re
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

RAW_DIR = Path(__file__).resolve().parents[1] / "data" / "raw"
OUT = Path(__file__).resolve().parents[1] / "data" / "struktura_pregled.csv"

DATE_RE = re.compile(r"na-dan-(\d{2})\.(\d{2})\.(\d{4})\.xlsx$")


def report_date(name: str) -> str:
    m = DATE_RE.search(name)
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else "?"


def main() -> None:
    rows = []
    header_variants: Counter = Counter()
    header_map: dict[tuple, list[str]] = {}

    files = sorted(RAW_DIR.glob("*.xlsx"), key=lambda p: report_date(p.name))
    for i, f in enumerate(files, 1):
        wb = openpyxl.load_workbook(f, read_only=True)
        sheets = wb.sheetnames
        rec = {
            "datum": report_date(f.name),
            "datoteka": f.name,
            "listi": "|".join(sheets),
            "tb01_stolpcev": None,
            "tb01_vrstic": None,
            "tb01_glava_id": None,
        }
        if "Tb 01" in sheets:
            ws = wb["Tb 01"]
            header = None
            n_data = 0
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = tuple(str(c).strip() if c is not None else "" for c in row)
                elif row[0] is not None:
                    n_data += 1
            header_variants[header] += 1
            header_map.setdefault(header, []).append(rec["datum"])
            rec["tb01_stolpcev"] = len(header)
            rec["tb01_vrstic"] = n_data
            rec["tb01_glava_id"] = list(header_variants.keys()).index(header)
        wb.close()
        rows.append(rec)
        if i % 20 == 0:
            print(f"  ... {i}/{len(files)}")

    df = pd.DataFrame(rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT, index=False, encoding="utf-8-sig")

    print(f"\nDatotek: {len(df)}")
    print("\nRazlicice seznamov listov:")
    for listi, n in Counter(df["listi"]).items():
        print(f"  {n:3d}x  {listi}")
    print(f"\nRazlicice glave 'Tb 01': {len(header_variants)}")
    for h, n in header_variants.items():
        dates = header_map[h]
        print(f"\n  {n:3d} datotek, {len(h)} stolpcev, od {min(dates)} do {max(dates)}")
        for c in h:
            print(f"      | {c}")
    print("\nRazpon stevila VZS (vrstic Tb 01):",
          df["tb01_vrstic"].min(), "-", df["tb01_vrstic"].max())


if __name__ == "__main__":
    main()
